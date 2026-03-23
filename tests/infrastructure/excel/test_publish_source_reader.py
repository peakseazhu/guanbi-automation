from pathlib import Path

from openpyxl import Workbook

from guanbi_automation.domain.publish_contract import PublishSourceSpec
from guanbi_automation.infrastructure.excel import publish_source_reader
from guanbi_automation.infrastructure.excel.publish_source_reader import read_publish_source


def test_read_sheet_source_excludes_header_and_trims_tail_blanks(tmp_path: Path):
    workbook_path = tmp_path / "result.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "计算表1"
    sheet["A1"] = "表头1"
    sheet["B1"] = "表头2"
    sheet["A2"] = "x"
    sheet["B2"] = 1
    sheet["A3"] = "y"
    sheet["B3"] = 2
    sheet["A4"] = None
    sheet["B4"] = None
    workbook.save(workbook_path)

    source = PublishSourceSpec(
        source_id="calc-1",
        sheet_name="计算表1",
        read_mode="sheet",
        start_row=1,
        start_col=1,
        header_mode="exclude",
    )

    dataset = read_publish_source(workbook_path, source)

    assert dataset.rows == [["x", 1], ["y", 2]]
    assert dataset.row_count == 2
    assert dataset.column_count == 2
    assert dataset.source_range == "计算表1!A2:B3"

    workbook_path.unlink()
    assert not workbook_path.exists()


def test_read_block_source_respects_declared_bounds(tmp_path: Path):
    workbook_path = tmp_path / "result.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "计算表2"
    sheet["A1"] = "忽略"
    sheet["B2"] = "区域"
    sheet["C2"] = "值"
    sheet["B3"] = "华东"
    sheet["C3"] = 88
    sheet["D3"] = "块外"
    workbook.save(workbook_path)

    source = PublishSourceSpec(
        source_id="calc-2",
        sheet_name="计算表2",
        read_mode="block",
        start_row=2,
        start_col=2,
        end_row=3,
        end_col=3,
        header_mode="include",
    )

    dataset = read_publish_source(workbook_path, source)

    assert dataset.rows == [["区域", "值"], ["华东", 88]]
    assert dataset.row_count == 2
    assert dataset.column_count == 2
    assert dataset.source_range == "计算表2!B2:C3"


def test_read_publish_source_opens_large_workbook_in_read_only_mode(
    tmp_path: Path,
    monkeypatch,
):
    workbook_path = tmp_path / "large-result.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "全国执行"
    sheet["A1"] = "日期"
    sheet["B1"] = "门店数"
    sheet["A2"] = "2026-03-22"
    sheet["B2"] = 3
    workbook.save(workbook_path)

    observed: dict[str, object] = {}
    original_load_workbook = publish_source_reader.load_workbook

    def tracking_load_workbook(*args, **kwargs):
        observed["data_only"] = kwargs.get("data_only")
        observed["read_only"] = kwargs.get("read_only")
        return original_load_workbook(*args, **kwargs)

    monkeypatch.setattr(publish_source_reader, "load_workbook", tracking_load_workbook)

    source = PublishSourceSpec(
        source_id="large-sheet",
        sheet_name="全国执行",
        read_mode="sheet",
        start_row=1,
        start_col=1,
        header_mode="include",
    )

    dataset = read_publish_source(workbook_path, source)

    assert observed == {"data_only": True, "read_only": True}
    assert dataset.rows == [["日期", "门店数"], ["2026-03-22", 3]]
