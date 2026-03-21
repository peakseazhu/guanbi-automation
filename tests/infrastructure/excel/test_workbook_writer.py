from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from guanbi_automation.domain.workbook_contract import WorkbookBlockSpec
from guanbi_automation.infrastructure.excel.workbook_writer import write_block



def test_replace_range_clears_values_but_preserves_existing_formula_cells(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    result_path = tmp_path / "result.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "底表"
    sheet["B2"] = 10
    sheet["B3"] = 20
    sheet["B4"] = 30
    sheet["B5"] = 40
    sheet["C2"] = "=B2*2"
    sheet["C3"] = "=B3*2"
    sheet["C4"] = "=B4*2"
    sheet["C5"] = "=B5*2"
    workbook.save(template_path)

    block = WorkbookBlockSpec.model_validate(
        {
            "block_id": "inventory_replace",
            "sheet_name": "底表",
            "source_extract_id": "inventory_daily",
            "write_mode": "replace_range",
            "start_row": 2,
            "start_col": 2,
            "end_row": 5,
            "end_col": 3,
            "clear_policy": "clear_values",
        }
    )

    result = write_block(
        template_path=template_path,
        result_path=result_path,
        block=block,
        rows=[[101], [102], [103]],
    )

    saved = load_workbook(result_path, data_only=False)
    saved_sheet = saved["底表"]

    assert result.written_start_row == 2
    assert result.written_end_row == 4
    assert saved_sheet["B2"].value == 101
    assert saved_sheet["B3"].value == 102
    assert saved_sheet["B4"].value == 103
    assert saved_sheet["B5"].value is None
    assert saved_sheet["C2"].value == "=B2*2"
    assert saved_sheet["C5"].value == "=B5*2"



def test_append_rows_starts_after_last_locator_row(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    result_path = tmp_path / "result.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "底表"
    sheet["B2"] = "门店A"
    sheet["C2"] = 100
    sheet["B3"] = "门店B"
    sheet["C3"] = 200
    sheet["B4"] = "门店C"
    sheet["C4"] = 300
    workbook.save(template_path)

    block = WorkbookBlockSpec.model_validate(
        {
            "block_id": "sales_append",
            "sheet_name": "底表",
            "source_extract_id": "sales_detail",
            "write_mode": "append_rows",
            "start_row": 2,
            "start_col": 2,
            "append_locator_columns": [2, 3],
            "clear_policy": "none",
        }
    )

    result = write_block(
        template_path=template_path,
        result_path=result_path,
        block=block,
        rows=[["门店D", 400], ["门店E", 500]],
    )

    saved = load_workbook(result_path, data_only=False)
    saved_sheet = saved["底表"]

    assert result.written_start_row == 5
    assert result.written_end_row == 6
    assert saved_sheet["B5"].value == "门店D"
    assert saved_sheet["C6"].value == 500



def test_fill_fixed_value_only_populates_newly_written_rows(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    result_path = tmp_path / "result.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "底表"
    sheet["B2"] = "门店A"
    sheet["C2"] = 100
    sheet["B3"] = "门店B"
    sheet["C3"] = 200
    sheet["B4"] = "门店C"
    sheet["C4"] = 300
    workbook.save(template_path)

    block = WorkbookBlockSpec.model_validate(
        {
            "block_id": "sales_append",
            "sheet_name": "底表",
            "source_extract_id": "sales_detail",
            "write_mode": "append_rows",
            "start_row": 2,
            "start_col": 2,
            "append_locator_columns": [2, 3],
            "clear_policy": "none",
            "post_write_actions": [
                {"action": "fill_fixed_value", "column": 1, "value": "华东区"}
            ],
        }
    )

    result = write_block(
        template_path=template_path,
        result_path=result_path,
        block=block,
        rows=[["门店D", 400], ["门店E", 500]],
    )

    saved = load_workbook(result_path, data_only=False)
    saved_sheet = saved["底表"]

    assert result.written_end_row == 6
    assert saved_sheet["A4"].value is None
    assert saved_sheet["A5"].value == "华东区"
    assert saved_sheet["A6"].value == "华东区"



def test_fill_down_formula_extends_formula_to_written_end_row(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    result_path = tmp_path / "result.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "底表"
    sheet["B2"] = "门店A"
    sheet["C2"] = 100
    sheet["D2"] = "=C2*2"
    sheet["B3"] = "门店B"
    sheet["C3"] = 200
    sheet["D3"] = "=C3*2"
    sheet["B4"] = "门店C"
    sheet["C4"] = 300
    sheet["D4"] = "=C4*2"
    workbook.save(template_path)

    block = WorkbookBlockSpec.model_validate(
        {
            "block_id": "sales_append",
            "sheet_name": "底表",
            "source_extract_id": "sales_detail",
            "write_mode": "append_rows",
            "start_row": 2,
            "start_col": 2,
            "append_locator_columns": [2, 3],
            "clear_policy": "none",
            "post_write_actions": [
                {"action": "fill_down_formula", "columns": [4]}
            ],
        }
    )

    result = write_block(
        template_path=template_path,
        result_path=result_path,
        block=block,
        rows=[["门店D", 400], ["门店E", 500]],
    )

    saved = load_workbook(result_path, data_only=False)
    saved_sheet = saved["底表"]

    assert result.actions["fill_down_formula"]["covered_end_row"] == 6
    assert saved_sheet["D5"].value == "=C5*2"
    assert saved_sheet["D6"].value == "=C6*2"



def test_fill_down_formula_fails_when_seed_formula_is_missing(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    result_path = tmp_path / "result.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "底表"
    sheet["B2"] = "门店A"
    sheet["C2"] = 100
    workbook.save(template_path)

    block = WorkbookBlockSpec.model_validate(
        {
            "block_id": "sales_append",
            "sheet_name": "底表",
            "source_extract_id": "sales_detail",
            "write_mode": "append_rows",
            "start_row": 2,
            "start_col": 2,
            "append_locator_columns": [2, 3],
            "clear_policy": "none",
            "post_write_actions": [
                {"action": "fill_down_formula", "columns": [4]}
            ],
        }
    )

    with pytest.raises(ValueError):
        write_block(
            template_path=template_path,
            result_path=result_path,
            block=block,
            rows=[["门店D", 400], ["门店E", 500]],
        )
