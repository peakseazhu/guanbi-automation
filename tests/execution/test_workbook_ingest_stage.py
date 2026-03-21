from pathlib import Path

from openpyxl import Workbook

from guanbi_automation.application.preflight_service import run_stage_preflight
from guanbi_automation.bootstrap.settings import WorkbookSettings
from guanbi_automation.domain.workbook_contract import WorkbookStageSpec
from guanbi_automation.execution.stages.workbook_ingest import (
    PlannedWorkbookIngestRun,
    WorkbookIngestStage,
)



def test_workbook_ingest_manifest_records_written_rows_and_actions(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    result_path = tmp_path / "result.xlsx"
    extract_path = tmp_path / "sales.xlsx"

    template = Workbook()
    template_sheet = template.active
    template_sheet.title = "底表"
    template_sheet["B2"] = "门店A"
    template_sheet["C2"] = 100
    template_sheet["B3"] = "门店B"
    template_sheet["C3"] = 200
    template_sheet["B4"] = "门店C"
    template_sheet["C4"] = 300
    template.save(template_path)

    extract = Workbook()
    extract_sheet = extract.active
    extract_sheet.title = "导出数据"
    extract_sheet.append(["门店D", 400])
    extract_sheet.append(["门店E", 500])
    extract.save(extract_path)

    stage_spec = WorkbookStageSpec.model_validate(
        {
            "template_path": str(template_path),
            "blocks": [
                {
                    "block_id": "sales_append",
                    "sheet_name": "底表",
                    "source_extract_id": "sales_detail",
                    "write_mode": "append_rows",
                    "start_row": 2,
                    "start_col": 2,
                    "append_locator_columns": [2, 3],
                    "clear_policy": "none",
                    "post_write_actions": [
                        {"action": "fill_fixed_value", "column": 1, "value": "华东区"}
                    ],
                }
            ],
        }
    )

    stage = WorkbookIngestStage()
    result = stage.run(
        PlannedWorkbookIngestRun(
            batch_id="batch-001",
            job_id="job-001",
            workbook_spec=stage_spec,
            result_path=result_path,
            extract_artifacts={"sales_detail": extract_path},
            workbook_settings=WorkbookSettings(),
        )
    )

    assert result.status == "completed"
    assert result.manifest["blocks"][0]["written_end_row"] == 6
    assert result.manifest["blocks"][0]["actions"]["fill_fixed_value"]["covered_end_row"] == 6



def test_preflight_delegates_to_workbook_gate(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    Workbook().save(template_path)

    report = run_stage_preflight(
        stage_name="workbook",
        row_count=10,
        column_count=5,
        cell_limit=5_000_000,
        template_path=template_path,
    )

    assert report.status == "ready"
