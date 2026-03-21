from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from shutil import copyfile

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from guanbi_automation.domain.workbook_contract import WorkbookBlockSpec, WorkbookPostWriteAction
from guanbi_automation.infrastructure.excel.block_locator import find_append_start_row


@dataclass(frozen=True)
class WriteBlockResult:
    workbook_path: Path
    written_start_row: int
    written_end_row: int
    written_start_col: int
    written_end_col: int
    actions: dict[str, object] = field(default_factory=dict)



def write_block(
    *,
    template_path: Path,
    result_path: Path,
    block: WorkbookBlockSpec,
    rows: list[list[object]],
) -> WriteBlockResult:
    template = Path(template_path)
    result = Path(result_path)
    if not result.exists():
        copyfile(template, result)

    workbook = load_workbook(result)
    sheet = workbook[block.sheet_name]

    source_width = max((len(row) for row in rows), default=0)
    if block.write_mode == "append_rows":
        write_start_row = _resolve_append_start_row(sheet=sheet, block=block)
    else:
        write_start_row = block.start_row

    write_start_col = block.start_col
    write_end_row = write_start_row + len(rows) - 1
    write_end_col = write_start_col + source_width - 1 if source_width else write_start_col

    if block.write_mode == "replace_sheet":
        _clear_values(
            sheet,
            start_row=block.start_row,
            start_col=block.start_col,
            end_row=max(sheet.max_row, write_end_row),
            end_col=max(sheet.max_column, write_end_col),
        )
    elif block.write_mode == "replace_range" and block.clear_policy == "clear_values":
        _clear_values(
            sheet,
            start_row=block.start_row,
            start_col=block.start_col,
            end_row=block.end_row or write_end_row,
            end_col=block.end_col or write_end_col,
        )

    _write_rows(sheet, rows=rows, start_row=write_start_row, start_col=write_start_col)
    action_results = _apply_post_write_actions(
        sheet=sheet,
        actions=block.post_write_actions,
        written_start_row=write_start_row,
        written_end_row=write_end_row,
    )
    workbook.save(result)

    return WriteBlockResult(
        workbook_path=result,
        written_start_row=write_start_row,
        written_end_row=write_end_row,
        written_start_col=write_start_col,
        written_end_col=write_end_col,
        actions=action_results,
    )



def _resolve_append_start_row(*, sheet: Worksheet, block: WorkbookBlockSpec) -> int:
    max_locator_col = max(block.append_locator_columns)
    rows: list[list[object]] = []
    for row_index in range(block.start_row, sheet.max_row + 1):
        rows.append(
            [sheet.cell(row=row_index, column=column).value for column in range(1, max_locator_col + 1)]
        )

    return find_append_start_row(
        rows=rows,
        anchor_row=block.start_row,
        locator_columns=block.append_locator_columns,
    )



def _clear_values(
    sheet: Worksheet,
    *,
    start_row: int,
    start_col: int,
    end_row: int,
    end_col: int,
) -> None:
    for row in range(start_row, end_row + 1):
        for column in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=column)
            if cell.data_type == "f":
                continue
            cell.value = None



def _write_rows(
    sheet: Worksheet,
    *,
    rows: list[list[object]],
    start_row: int,
    start_col: int,
) -> None:
    for row_offset, row in enumerate(rows):
        for column_offset, value in enumerate(row):
            sheet.cell(
                row=start_row + row_offset,
                column=start_col + column_offset,
                value=value,
            )



def _apply_post_write_actions(
    *,
    sheet: Worksheet,
    actions: list[WorkbookPostWriteAction],
    written_start_row: int,
    written_end_row: int,
) -> dict[str, object]:
    results: dict[str, object] = {}

    for action in actions:
        if action.action == "fill_fixed_value":
            if action.column is None:
                raise ValueError("fill_fixed_value requires a target column")
            for row in range(written_start_row, written_end_row + 1):
                sheet.cell(row=row, column=action.column, value=action.value)
            results[action.action] = {
                "column": action.column,
                "covered_end_row": written_end_row,
            }
        elif action.action == "fill_down_formula":
            covered_columns: list[int] = []
            for column in action.columns:
                seed_row, seed_formula = _find_seed_formula(
                    sheet=sheet,
                    column=column,
                    written_start_row=written_start_row,
                )
                origin = f"{get_column_letter(column)}{seed_row}"
                for row in range(written_start_row, written_end_row + 1):
                    target = f"{get_column_letter(column)}{row}"
                    translated = Translator(seed_formula, origin=origin).translate_formula(target)
                    sheet[target] = translated
                covered_columns.append(column)
            results[action.action] = {
                "columns": covered_columns,
                "covered_end_row": written_end_row,
            }

    return results



def _find_seed_formula(
    *,
    sheet: Worksheet,
    column: int,
    written_start_row: int,
) -> tuple[int, str]:
    for row in range(written_start_row - 1, 0, -1):
        cell = sheet.cell(row=row, column=column)
        if cell.data_type == "f" and isinstance(cell.value, str):
            return row, cell.value

    raise ValueError(f"Missing seed formula for column {column}")
