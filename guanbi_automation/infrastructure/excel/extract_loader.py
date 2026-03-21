from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from guanbi_automation.infrastructure.excel.block_locator import trim_trailing_empty_edges


@dataclass(frozen=True)
class ExtractTable:
    rows: list[list[object]]
    row_count: int
    column_count: int
    cell_count: int


def load_extract_table(path: Path) -> ExtractTable:
    file_path = Path(path)
    suffix = file_path.suffix.lower()

    if suffix == ".xlsx":
        rows = _load_xlsx_rows(file_path)
    elif suffix == ".csv":
        rows = _load_csv_rows(file_path)
    else:
        raise ValueError(f"Unsupported extract artifact: {file_path.suffix}")

    normalized_rows = trim_trailing_empty_edges(rows)
    row_count = len(normalized_rows)
    column_count = max((len(row) for row in normalized_rows), default=0)
    cell_count = row_count * column_count
    return ExtractTable(
        rows=normalized_rows,
        row_count=row_count,
        column_count=column_count,
        cell_count=cell_count,
    )


def _load_xlsx_rows(path: Path) -> list[list[object]]:
    workbook = load_workbook(path, data_only=False)
    sheet = workbook.worksheets[0]
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def _load_csv_rows(path: Path) -> list[list[object]]:
    with path.open("r", newline="", encoding="utf-8") as handle:
        return [list(row) for row in csv.reader(handle)]
