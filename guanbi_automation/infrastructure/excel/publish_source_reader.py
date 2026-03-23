from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from guanbi_automation.domain.publish_contract import PublishDataset, PublishSourceSpec
from guanbi_automation.infrastructure.excel.block_locator import trim_trailing_empty_edges


def read_publish_source(workbook_path: Path, source: PublishSourceSpec) -> PublishDataset:
    # Live verification uses a 215MB workbook sample, so publish-source reads must stay streaming-safe.
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        sheet = workbook[source.sheet_name]
        rows = _read_bounded_rows(sheet=sheet, source=source)
        start_row = source.start_row

        if source.header_mode == "exclude" and rows:
            rows = rows[1:]
            start_row += 1

        trimmed_rows = trim_trailing_empty_edges(rows)
        row_count = len(trimmed_rows)
        column_count = max((len(row) for row in trimmed_rows), default=0)
        return PublishDataset(
            rows=trimmed_rows,
            row_count=row_count,
            column_count=column_count,
            source_range=_format_source_range(
                sheet_name=source.sheet_name,
                start_row=start_row,
                start_col=source.start_col,
                rows=trimmed_rows,
            ),
        )
    finally:
        workbook.close()


def _read_bounded_rows(*, sheet: object, source: PublishSourceSpec) -> list[list[object]]:
    max_row = source.end_row or getattr(sheet, "max_row")
    max_col = source.end_col or getattr(sheet, "max_column")
    if max_row is None or max_col is None:
        return []

    rows = [
        list(row)
        for row in sheet.iter_rows(
            min_row=source.start_row,
            max_row=max_row,
            min_col=source.start_col,
            max_col=max_col,
            values_only=True,
        )
    ]

    return trim_trailing_empty_edges(rows)


def _format_source_range(
    *,
    sheet_name: str,
    start_row: int,
    start_col: int,
    rows: list[list[object]],
) -> str:
    if not rows:
        return (
            f"{sheet_name}!"
            f"{_column_label(start_col)}{start_row}:"
            f"{_column_label(start_col)}{start_row}"
        )

    end_row = start_row + len(rows) - 1
    end_col = start_col + max((len(row) for row in rows), default=1) - 1
    return (
        f"{sheet_name}!"
        f"{_column_label(start_col)}{start_row}:"
        f"{_column_label(end_col)}{end_row}"
    )


def _column_label(column_number: int) -> str:
    label = ""
    current = column_number

    while current > 0:
        current, remainder = divmod(current - 1, 26)
        label = chr(65 + remainder) + label

    return label
